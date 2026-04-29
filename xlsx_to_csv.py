#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

FORMULA_WARNING = (
    "Formula cells detected. CSV contains cached values stored by Excel; "
    "formulas were not recalculated."
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export selected XLSX worksheet data to UTF-8-SIG CSV files."
    )
    parser.add_argument("input_file", type=Path, help="Path to source .xlsx workbook")
    parser.add_argument("--outdir", type=Path, default=Path("."), help="Output directory")
    parser.add_argument("--all-sheets", action="store_true", help="Export all visible sheets")
    parser.add_argument(
        "--sheets",
        type=str,
        help="Comma-separated list of sheet names to export (exact names)",
    )
    parser.add_argument("--list-sheets", action="store_true", help="List sheet names only")
    parser.add_argument("--force", action="store_true", help="Overwrite existing CSV outputs")
    parser.add_argument(
        "--include-hidden-sheets",
        action="store_true",
        help="Include hidden and veryHidden sheets in export selection",
    )
    parser.add_argument("--header-row", type=int, help="1-indexed header row number")
    parser.add_argument(
        "--normalize-headers",
        action="store_true",
        help="Normalize header names and deduplicate with suffixes",
    )
    args = parser.parse_args()

    if args.header_row is not None and args.header_row <= 0:
        parser.error("--header-row must be >= 1")

    if args.all_sheets and args.sheets:
        parser.error("Use only one of --all-sheets or --sheets")

    return args


def sanitize_filename_part(value: str) -> str:
    text = re.sub(r"[\s]+", "_", value.strip())
    text = re.sub(r"[/\\:*?\"<>|]", "_", text)
    text = re.sub(r"_+", "_", text)
    text = text.strip("_")
    return text or "unnamed"


def normalize_header(value: str) -> str:
    text = value.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text)
    text = text.strip("_")
    return text or "column"


def make_unique_headers(headers: list[str]) -> tuple[list[str], bool]:
    counts: dict[str, int] = {}
    unique: list[str] = []
    had_duplicates = False
    for header in headers:
        base = header or "column"
        current = counts.get(base, 0) + 1
        counts[base] = current
        if current == 1:
            unique.append(base)
        else:
            had_duplicates = True
            unique.append(f"{base}_{current}")
    return unique, had_duplicates


def list_sheets(workbook_path: Path) -> list[str]:
    wb = load_workbook(filename=workbook_path, read_only=True, data_only=False)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def detect_used_range(sheet: Worksheet) -> tuple[int | None, int | None, int]:
    max_row_seen = 0
    max_col_seen = 0
    first_nonempty_row: int | None = None
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        row_has_data = False
        row_last_col = 0
        for c_idx, val in enumerate(row, start=1):
            if val is not None:
                row_has_data = True
                row_last_col = c_idx
        if row_has_data:
            if first_nonempty_row is None:
                first_nonempty_row = r_idx
            max_row_seen = r_idx
            if row_last_col > max_col_seen:
                max_col_seen = row_last_col
    return first_nonempty_row, (max_row_seen or None), max_col_seen


def convert_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value)


def inspect_workbook_structure(workbook_path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(filename=workbook_path, data_only=False)
    inspections: dict[str, dict[str, Any]] = {}
    try:
        for ws in wb.worksheets:
            formula_cell_count = 0
            comments_count = 0
            hidden_rows_count = sum(1 for d in ws.row_dimensions.values() if d.hidden)
            hidden_columns_count = sum(1 for d in ws.column_dimensions.values() if d.hidden)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        formula_cell_count += 1
                    if cell.comment is not None:
                        comments_count += 1

            data_validations_count = len(ws.data_validations.dataValidation) if ws.data_validations else 0
            table_count = len(ws.tables)
            inspections[ws.title] = {
                "sheet_state": ws.sheet_state,
                "formula_cell_count": formula_cell_count,
                "merged_cell_ranges_count": len(ws.merged_cells.ranges),
                "hidden_rows_count": hidden_rows_count,
                "hidden_columns_count": hidden_columns_count,
                "comments_count": comments_count,
                "data_validations_count": data_validations_count,
                "table_count": table_count,
            }
    finally:
        wb.close()
    return inspections


def load_workbook_for_values(workbook_path: Path):
    return load_workbook(filename=workbook_path, data_only=True)


def export_sheet_to_csv(
    *,
    ws: Worksheet,
    output_csv: Path,
    force: bool,
    header_row: int | None,
    normalize_headers_flag: bool,
) -> dict[str, Any]:
    details: dict[str, Any] = {
        "exported": False,
        "skipped_reason": None,
        "output_csv": str(output_csv),
        "row_count": 0,
        "column_count": 0,
        "header_row_used": None,
        "header_values_original": [],
        "header_values_normalized": None,
        "duplicate_headers_detected": False,
        "warnings": [],
    }

    first_nonempty_row, max_row, max_col = detect_used_range(ws)
    if first_nonempty_row is None or max_row is None or max_col == 0:
        details["skipped_reason"] = "empty_sheet"
        details["warnings"].append("Sheet is empty and was skipped.")
        return details

    header_row_used = header_row if header_row is not None else first_nonempty_row
    details["header_row_used"] = header_row_used

    if output_csv.exists() and not force:
        details["skipped_reason"] = "output_exists"
        return details

    output_csv.parent.mkdir(parents=True, exist_ok=True)

    rows_for_export: list[list[str]] = []
    for r_idx in range(1, max_row + 1):
        row_values: list[str] = []
        for c_idx in range(1, max_col + 1):
            row_values.append(convert_cell_value(ws.cell(row=r_idx, column=c_idx).value))
        rows_for_export.append(row_values)

    header_values_original = rows_for_export[header_row_used - 1] if 1 <= header_row_used <= len(rows_for_export) else []
    details["header_values_original"] = header_values_original

    if normalize_headers_flag and header_values_original:
        normalized = [normalize_header(v) for v in header_values_original]
        unique_headers, dupes = make_unique_headers(normalized)
        rows_for_export[header_row_used - 1] = unique_headers
        details["header_values_normalized"] = unique_headers
        details["duplicate_headers_detected"] = dupes or (len(normalized) != len(set(normalized)))
    elif header_values_original:
        raw = [v.strip() for v in header_values_original]
        counts = Counter(raw)
        details["duplicate_headers_detected"] = any(v > 1 and k != "" for k, v in counts.items())

    with output_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle, quoting=csv.QUOTE_MINIMAL)
        writer.writerows(rows_for_export)

    details["exported"] = True
    details["row_count"] = len(rows_for_export)
    details["column_count"] = max_col
    return details


def write_report(report_path: Path, report_data: dict[str, Any]) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report_data, indent=2, ensure_ascii=False), encoding="utf-8")


def main() -> int:
    args = parse_args()
    report_warnings: list[str] = []

    input_file = args.input_file
    report: dict[str, Any] = {
        "report_schema_version": "1.0",
        "input_file": str(input_file),
        "output_dir": str(args.outdir),
        "openpyxl_version": openpyxl.__version__,
        "export_status": "failed",
        "workbook_size_bytes": None,
        "sheets_found": [],
        "visible_sheets": [],
        "hidden_sheets": [],
        "selected_sheets": [],
        "exported_sheets": [],
        "skipped_sheets": [],
        "total_csv_files_written": 0,
        "total_rows_written": 0,
        "total_columns_written_max": 0,
        "warnings": report_warnings,
        "sheets": [],
    }

    if not input_file.exists():
        print(f"Error: input file does not exist: {input_file}", file=sys.stderr)
        report_warnings.append("Input file does not exist.")
        return 1
    if input_file.suffix.lower() != ".xlsx":
        print("Error: input file must be a .xlsx file", file=sys.stderr)
        report_warnings.append("Input file is not .xlsx.")
        return 1

    report["workbook_size_bytes"] = input_file.stat().st_size

    try:
        inspections = inspect_workbook_structure(input_file)
        wb_values = load_workbook_for_values(input_file)
    except Exception as exc:
        print(f"Error: workbook could not be opened: {exc}", file=sys.stderr)
        report_warnings.append("Workbook cannot be opened.")
        return 1

    try:
        all_sheet_names = wb_values.sheetnames
        report["sheets_found"] = all_sheet_names
        visible = [ws.title for ws in wb_values.worksheets if ws.sheet_state == "visible"]
        hidden = [ws.title for ws in wb_values.worksheets if ws.sheet_state != "visible"]
        report["visible_sheets"] = visible
        report["hidden_sheets"] = hidden

        if args.list_sheets:
            for name in all_sheet_names:
                print(name)
            report["export_status"] = "success"
            return 0

        if args.sheets:
            selected_sheets = [s.strip() for s in args.sheets.split(",") if s.strip()]
        elif args.all_sheets:
            selected_sheets = visible if not args.include_hidden_sheets else all_sheet_names
        else:
            selected_sheets = [wb_values.active.title]
            report_warnings.append(
                "Neither --sheets nor --all-sheets supplied; exported active sheet only."
            )

        missing = [s for s in selected_sheets if s not in all_sheet_names]
        if missing:
            print(f"Error: selected sheets not found: {', '.join(missing)}", file=sys.stderr)
            report_warnings.append("Selected sheet does not exist.")
            return 1

        report["selected_sheets"] = selected_sheets

        workbook_stub = sanitize_filename_part(input_file.stem)
        had_export_error = False

        for ws in wb_values.worksheets:
            name = ws.title
            meta = inspections.get(name, {})
            sheet_state = meta.get("sheet_state", ws.sheet_state)
            is_selected = name in selected_sheets
            sheet_entry: dict[str, Any] = {
                "sheet_name": name,
                "sheet_state": sheet_state,
                "selected": is_selected,
                "exported": False,
                "skipped_reason": None,
                "output_csv": None,
                "row_count": 0,
                "column_count": 0,
                "header_row_used": None,
                "header_values_original": [],
                "header_values_normalized": None,
                "duplicate_headers_detected": False,
                "formula_cell_count": meta.get("formula_cell_count", 0),
                "merged_cell_ranges_count": meta.get("merged_cell_ranges_count", 0),
                "hidden_rows_count": meta.get("hidden_rows_count", 0),
                "hidden_columns_count": meta.get("hidden_columns_count", 0),
                "comments_count": meta.get("comments_count", 0),
                "warnings": [],
            }

            if not is_selected:
                report["sheets"].append(sheet_entry)
                continue

            if sheet_state != "visible" and not args.include_hidden_sheets:
                sheet_entry["skipped_reason"] = "hidden_sheet"
                report["skipped_sheets"].append(name)
                report["sheets"].append(sheet_entry)
                continue

            if meta.get("formula_cell_count", 0) > 0:
                sheet_entry["warnings"].append(FORMULA_WARNING)
                if FORMULA_WARNING not in report_warnings:
                    report_warnings.append(FORMULA_WARNING)

            if meta.get("merged_cell_ranges_count", 0) > 0:
                sheet_entry["warnings"].append("Merged cells detected; CSV cannot preserve merged structure.")
            if meta.get("comments_count", 0) > 0:
                sheet_entry["warnings"].append("Comments detected; CSV cannot preserve comments.")
            if meta.get("hidden_rows_count", 0) > 0:
                sheet_entry["warnings"].append("Hidden rows detected; CSV cannot preserve hidden-row state.")
            if meta.get("hidden_columns_count", 0) > 0:
                sheet_entry["warnings"].append("Hidden columns detected; CSV cannot preserve hidden-column state.")
            if meta.get("data_validations_count", 0) > 0:
                sheet_entry["warnings"].append("Data validations detected; CSV cannot preserve validations.")
            if meta.get("table_count", 0) > 1:
                sheet_entry["warnings"].append("Multiple tables detected; CSV flattens worksheet grid only.")

            sheet_csv_name = f"{workbook_stub}__{sanitize_filename_part(name)}.csv"
            output_csv = args.outdir / sheet_csv_name

            try:
                result = export_sheet_to_csv(
                    ws=ws,
                    output_csv=output_csv,
                    force=args.force,
                    header_row=args.header_row,
                    normalize_headers_flag=args.normalize_headers,
                )
            except Exception as exc:
                had_export_error = True
                sheet_entry["skipped_reason"] = f"export_error: {exc}"
                report["skipped_sheets"].append(name)
                report["sheets"].append(sheet_entry)
                continue

            for key, value in result.items():
                sheet_entry[key] = value

            if sheet_entry["exported"]:
                report["exported_sheets"].append(name)
                report["total_csv_files_written"] += 1
                report["total_rows_written"] += sheet_entry["row_count"]
                report["total_columns_written_max"] = max(
                    report["total_columns_written_max"], sheet_entry["column_count"]
                )
            else:
                report["skipped_sheets"].append(name)

            report["sheets"].append(sheet_entry)

        if had_export_error and report["total_csv_files_written"] > 0:
            report["export_status"] = "partial_success"
            return_code = 0
        elif had_export_error and report["total_csv_files_written"] == 0:
            report["export_status"] = "failed"
            return_code = 1
        elif report["selected_sheets"] and report["total_csv_files_written"] == 0:
            report["export_status"] = "partial_success"
            return_code = 0
        else:
            report["export_status"] = "success"
            return_code = 0

        write_report(args.outdir / "xlsx_export_report.json", report)
        return return_code
    finally:
        wb_values.close()


if __name__ == "__main__":
    raise SystemExit(main())
